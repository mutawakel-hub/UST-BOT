"""
إنشاء قالب Excel شامل لجمع بيانات المواد الدراسية
جامعة العلوم والتكنولوجيا - اليمن
=====================================
الأوراق (Sheets):
1. الإرشادات - تعليمات التعبئة
2. المواد - القالب الرئيسي للتعبئة
3. الكليات_والتخصصات - قائمة مرجعية معبأة مسبقاً
4. مثال_تقنية_المعلومات - مثال عملي للتعبئة الصحيحة
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ============================================
# الألوان (هوية جامعة العلوم والتكنولوجيا)
# ============================================
PRIMARY = "1B2A4A"        # أزرق داكن
PRIMARY_LIGHT = "D6E4F0"  # أزرق فاتح
ACCENT_POSITIVE = "1B7D46"  # أخضر
ACCENT_NEGATIVE = "C0392B"  # أحمر
ACCENT_WARNING = "D4820A"   # برتقالي

NEUTRAL_900 = "37352F"  # نص أساسي
NEUTRAL_600 = "8C8A84"  # نص ثانوي
NEUTRAL_200 = "E9E9E8"  # حدود
NEUTRAL_100 = "F7F7F5"  # خلفية الصفوف الفردية
NEUTRAL_0 = "FFFFFF"    # أبيض

# ============================================
# الخطوط
# ============================================
FONT_NAME = "Calibri"
HEADER_BOLD = True

# ============================================
# الأنماط
# ============================================
def title_style():
    return {
        "font": Font(name=FONT_NAME, size=18, bold=True, color=PRIMARY),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor=NEUTRAL_0),
    }

def subtitle_style():
    return {
        "font": Font(name=FONT_NAME, size=12, italic=True, color=NEUTRAL_600),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor=NEUTRAL_0),
    }

def section_header_style():
    return {
        "font": Font(name=FONT_NAME, size=12, bold=True, color=NEUTRAL_0),
        "alignment": Alignment(horizontal="right", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor=PRIMARY),
        "border": Border(
            bottom=Side(style="thin", color=NEUTRAL_200)
        ),
    }

def column_header_style():
    return {
        "font": Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_0),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor=PRIMARY),
        "border": Border(
            bottom=Side(style="medium", color=PRIMARY),
            top=Side(style="thin", color=PRIMARY),
        ),
    }

def data_cell_style(odd_row=True):
    fill_color = NEUTRAL_100 if odd_row else NEUTRAL_0
    return {
        "font": Font(name=FONT_NAME, size=11, color=NEUTRAL_900),
        "alignment": Alignment(horizontal="right", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor=fill_color),
        "border": Border(
            bottom=Side(style="thin", color=NEUTRAL_200)
        ),
    }

def instruction_step_style():
    return {
        "font": Font(name=FONT_NAME, size=11, color=NEUTRAL_900),
        "alignment": Alignment(horizontal="right", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor=NEUTRAL_0),
    }

def warning_style():
    return {
        "font": Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_NEGATIVE),
        "alignment": Alignment(horizontal="right", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor="FDEDEC"),
    }

def success_style():
    return {
        "font": Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_POSITIVE),
        "alignment": Alignment(horizontal="right", vertical="center", wrap_text=True),
        "fill": PatternFill("solid", fgColor="EAFAF1"),
    }

def apply_style(cell, style_dict):
    """تطبيق نمط على خلية"""
    if "font" in style_dict:
        cell.font = style_dict["font"]
    if "alignment" in style_dict:
        cell.alignment = style_dict["alignment"]
    if "fill" in style_dict:
        cell.fill = style_dict["fill"]
    if "border" in style_dict:
        cell.border = style_dict["border"]

# ============================================
# بيانات الكليات والتخصصات
# ============================================
COLLEGES = [
    (1, "كلية الطب والعلوم الصحية", "🏥"),
    (2, "كلية طب الأسنان", "🦷"),
    (3, "كلية الصيدلة", "💊"),
    (4, "كلية الهندسة", "⚙️"),
    (5, "كلية الحاسبات وتكنولوجيا المعلومات", "💻"),
    (6, "كلية العلوم الإدارية", "📊"),
    (7, "كلية العلوم الإنسانية والاجتماعية", "📚"),
]

SPECIALTIES = [
    (1, 1, "طب وجراحة", 6),
    (2, 1, "تكنولوجيا الأشعة التشخيصية", 4),
    (3, 1, "تغذية علاجية وحميات", 4),
    (4, 1, "الطب المخبري", 4),
    (5, 2, "طب وجراحة الفم والأسنان", 5),
    (6, 3, "دكتور صيدلة", 6),
    (7, 3, "صيدلة", 5),
    (8, 4, "هندسة الميكاترونكس", 5),
    (9, 4, "هندسة طبية حيوية", 5),
    (10, 4, "هندسة مدنية", 5),
    (11, 4, "هندسة معمارية - هندسة التصميم الداخلي", 5),
    (12, 4, "هندسة الحاسوب والأنظمة الذكية", 5),
    (13, 4, "هندسة الاتصالات والمعلوماتية", 5),
    (14, 4, "هندسة الطاقة المتجددة والتحكم الآلي", 5),
    (15, 5, "تقنية معلومات باللغة الإنجليزية (BIT)", 4),
    (16, 5, "تقنية معلومات (IT)", 4),
    (17, 5, "جرافكس وإعلام رقمي", 4),
    (18, 5, "الذكاء الاصطناعي", 4),
    (19, 5, "الأمن السيبراني والشبكات", 4),
    (20, 5, "هندسة البرمجيات", 4),
    (21, 5, "أعمال إلكترونية", 4),
    (22, 5, "ذكاء الأعمال - نظم المعلومات الإدارية", 4),
    (23, 6, "إدارة أعمال باللغة الإنجليزية", 4),
    (24, 6, "إدارة أعمال دولية - إدارة أعمال", 4),
    (25, 6, "التسويق الرقمي", 4),
    (26, 6, "محاسبة - علوم مالية ومصرفية", 4),
    (27, 6, "إدارة أعمال دولية باللغة الإنجليزية", 4),
    (28, 7, "لغة إنجليزية - ترجمة", 4),
    (29, 7, "لغة إنجليزية - لغويات تطبيقية", 4),
    (30, 7, "العلاقات العامة والإعلان", 4),
    (31, 7, "إذاعة وتلفزيون", 4),
    (32, 7, "علم النفس", 4),
    (33, 7, "شريعة وقانون", 4),
    (34, 7, "دراسات إسلامية - لغة عربية - علوم قرآن", 4),
]

# مواد IT كمثال
IT_SUBJECTS_EXAMPLE = [
    (1, 1, "مقدمة في تقنية المعلومات", "نعم", "نعم"),
    (1, 1, "برمجة حاسوب (1) - Python", "نعم", "نعم"),
    (1, 1, "الرياضيات المتقطعة", "نعم", "لا"),
    (1, 1, "مهارات التعلم والاتصال", "نعم", "لا"),
    (1, 1, "اللغة الإنجليزية (1)", "نعم", "لا"),
    (1, 2, "برمجة حاسوب (2) - Java", "نعم", "نعم"),
    (1, 2, "تراكيب البيانات", "نعم", "نعم"),
    (1, 2, "قواعد البيانات (1)", "نعم", "نعم"),
    (1, 2, "نظم التشغيل (1)", "نعم", "لا"),
    (1, 2, "اللغة الإنجليزية (2)", "نعم", "لا"),
    (2, 1, "البرمجة الكائنية (OOP)", "نعم", "نعم"),
    (2, 1, "هياكل البيانات المتقدمة", "نعم", "نعم"),
    (2, 1, "قواعد البيانات (2)", "نعم", "نعم"),
    (2, 1, "شبكات الحاسوب (1)", "نعم", "نعم"),
    (2, 1, "اللغة الإنجليزية (3)", "نعم", "لا"),
    (2, 2, "هندسة البرمجيات", "نعم", "لا"),
    (2, 2, "تطوير الويب (Frontend)", "نعم", "نعم"),
    (2, 2, "الخوارزميات", "نعم", "نعم"),
    (2, 2, "أمن المعلومات", "نعم", "لا"),
    (2, 2, "اللغة الإنجليزية (4)", "نعم", "لا"),
]

# ============================================
# إنشاء الـ Workbook
# ============================================
wb = Workbook()
wb.properties.creator = "UST Central Bot"
wb.properties.title = "قالب المواد الدراسية - جامعة العلوم والتكنولوجيا"

# حذف الورقة الافتراضية
default_sheet = wb.active
wb.remove(default_sheet)

# ============================================
# الورقة 1: الإرشادات
# ============================================
ws1 = wb.create_sheet("📋 الإرشادات", 0)
ws1.sheet_view.rightToLeft = True

# ضبط عرض الأعمدة
ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 55
ws1.column_dimensions['D'].width = 3

# العنوان الرئيسي
ws1.merge_cells('B2:C2')
ws1['B2'] = "📋 قالب جمع بيانات المواد الدراسية"
apply_style(ws1['B2'], title_style())
ws1.row_dimensions[2].height = 35

ws1.merge_cells('B3:C3')
ws1['B3'] = "جامعة العلوم والتكنولوجيا - اليمن | البوت العلمي المركزي"
apply_style(ws1['B3'], subtitle_style())
ws1.row_dimensions[3].height = 22

# مقدمة
ws1.merge_cells('B5:C5')
ws1['B5'] = "📝 مقدمة"
apply_style(ws1['B5'], section_header_style())
ws1.row_dimensions[5].height = 28

ws1.merge_cells('B6:C8')
ws1['B6'] = (
    "هذا القالب مُخصّص لجمع بيانات المواد الدراسية لكل تخصص في الجامعة. "
    "سيتم استخدام هذه البيانات لاحقاً في بناء قاعدة بيانات البوت العلمي المركزي. "
    "يرجى تعبئة الورقة الثانية (📝 المواد) لكل مادة في كل تخصص وكل مستوى وكل فصل."
)
apply_style(ws1['B6'], instruction_step_style())
ws1.row_dimensions[6].height = 22

# خطوات التعبئة
ws1.merge_cells('B10:C10')
ws1['B10'] = "✅ خطوات التعبئة"
apply_style(ws1['B10'], section_header_style())
ws1.row_dimensions[10].height = 28

steps = [
    ("1️⃣", "افتح ورقة (📝 المواد)", "هي الورقة الرئيسية للتعبئة. كل سطر يمثّل مادة واحدة."),
    ("2️⃣", "راجع ورقة (🏛 الكليات_والتخصصات)", "تحتوي على قائمة الكليات الـ 7 والتخصصات الـ 34. استخدم الأسماء كما هي تماماً."),
    ("3️⃣", "لكل مادة، املأ صفّاً جديداً", "لا تترك أي صف فارغ بين المواد. ابدأ من الصف 4 مباشرة."),
    ("4️⃣", "الكليات والتخصصات", "اختر من القائمة المنسدلة (سيظهر سهم صغير عند الضغط على الخلية)."),
    ("5️⃣", "المستوى (1-6)", "اكتب رقم المستوى. لكل تخصص عدد مستويات محدد (انظر العمود D في ورقة الكليات)."),
    ("6️⃣", "الفصل الدراسي", "اختر من القائمة المنسدلة: (1) للفصل الأول أو (2) للفصل الثاني."),
    ("7️⃣", "اسم المادة", "اكتب الاسم الكامل بالعربية. مثال: «برمجة حاسوب (1) - Python»"),
    ("8️⃣", "له مقرر نظري؟", "اختر من القائمة: نعم / لا"),
    ("9️⃣", "له مقرر عملي؟", "اختر من القائمة: نعم / لا"),
    ("🔟", "احفظ الملف", "احفظ باسم: «المواد - اسم الكلية.xlsx» وأرسله للمسؤول المركزي."),
]

for i, (num, title, desc) in enumerate(steps):
    row = 11 + i
    ws1.cell(row=row, column=2, value=f"{num} {title}")
    apply_style(ws1.cell(row=row, column=2), instruction_step_style())
    ws1.cell(row=row, column=3, value=desc)
    apply_style(ws1.cell(row=row, column=3), instruction_step_style())
    ws1.row_dimensions[row].height = 22

# تحذيرات
warning_row = 11 + len(steps) + 1
ws1.merge_cells(f'B{warning_row}:C{warning_row}')
ws1.cell(row=warning_row, column=2, value="⚠️ تحذيرات مهمة")
apply_style(ws1.cell(row=warning_row, column=2), section_header_style())
ws1.row_dimensions[warning_row].height = 28

warnings_list = [
    "لا تُعدّل أسماء الكليات أو التخصصات الموجودة في ورقة (🏛 الكليات_والتخصصات).",
    "لا تترك أي خلية فارغة في صف المادة. كل الخلايا مطلوبة.",
    "لا تكرّر نفس المادة في نفس الفصل (نفس التخصص + المستوى + الفصل + الاسم).",
    "تأكد من أن رقم المستوى لا يتجاوز عدد المستويات المحدد للتخصص.",
    "لا تضف أعمدة جديدة أو تُعدّل ترتيب الأعمدة الموجودة.",
    "إذا كان هناك مادة تدرّس في الفصلين، اكتبها في صفّين منفصلين (مرة في فصل 1، مرة في فصل 2).",
]

for i, w in enumerate(warnings_list):
    row = warning_row + 1 + i
    ws1.merge_cells(f'B{row}:C{row}')
    ws1.cell(row=row, column=2, value=f"⚠️ {w}")
    apply_style(ws1.cell(row=row, column=2), warning_style())
    ws1.row_dimensions[row].height = 22

# معلومات الاتصال
contact_row = warning_row + 1 + len(warnings_list) + 1
ws1.merge_cells(f'B{contact_row}:C{contact_row}')
ws1.cell(row=contact_row, column=2, value="📞 معلومات الاتصال")
apply_style(ws1.cell(row=contact_row, column=2), section_header_style())
ws1.row_dimensions[contact_row].height = 28

contacts = [
    ("للأسئلة والاستفسارات", "تواصل مع المسؤول المركزي عبر تيليجرام: @ust_support"),
    ("إرسال الملف المكتمل", "أرسل الملف بصيغة .xlsx إلى المسؤول المركزي"),
    ("الاستفسارات التقنية", "البريد: support@ust.edu.ye"),
]

for i, (label, value) in enumerate(contacts):
    row = contact_row + 1 + i
    ws1.cell(row=row, column=2, value=label)
    apply_style(ws1.cell(row=row, column=2), instruction_step_style())
    ws1.cell(row=row, column=3, value=value)
    apply_style(ws1.cell(row=row, column=3), instruction_step_style())
    ws1.row_dimensions[row].height = 22

# ============================================
# الورقة 2: المواد (القالب الرئيسي للتعبئة)
# ============================================
ws2 = wb.create_sheet("📝 المواد", 1)
ws2.sheet_view.rightToLeft = True

# العنوان
ws2.merge_cells('A1:H1')
ws2['A1'] = "📝 جدول المواد الدراسية - يُرجى التعبئة من السطر 4"
apply_style(ws2['A1'], title_style())
ws2.row_dimensions[1].height = 32

# ترتيب الأعمدة
headers = [
    ("A", "الكلية", 32),
    ("B", "التخصص", 38),
    ("C", "المستوى", 12),
    ("D", "الفصل", 12),
    ("E", "اسم المادة", 40),
    ("F", "له مقرر نظري؟", 18),
    ("G", "له مقرر عملي؟", 18),
    ("H", "ملاحظات (اختياري)", 25),
]

for col_letter, header, width in headers:
    ws2.column_dimensions[col_letter].width = width
    cell = ws2[f"{col_letter}3"]
    cell.value = header
    apply_style(cell, column_header_style())

ws2.row_dimensions[3].height = 35

# إضافة تعليقات للترويسة
ws2['A3'].comment = Comment(
    "اختر من القائمة المنسدلة.\n"
    "لا تكتب اسم الكلية يدوياً - استخدم القائمة فقط.",
    "نظام البوت"
)
ws2['B3'].comment = Comment(
    "اختر من القائمة المنسدلة.\n"
    "إذا لم يظهر التخصص المطلوب، تأكد من اختيار الكلية الصحيحة أولاً.",
    "نظام البوت"
)
ws2['C3'].comment = Comment(
    "اكتب رقم المستوى (1، 2، 3، ...).\n"
    "راجع الحد الأقصى لكل تخصص في ورقة الكليات_والتخصصات.",
    "نظام البوت"
)
ws2['E3'].comment = Comment(
    "اكتب الاسم الكامل بالعربية.\n"
    "مثال: «برمجة حاسوب (1) - Python»\n"
    "تجنّب الاختصارات غير الواضحة.",
    "نظام البوت"
)

# Data Validation: قائمة الكليات (العمود A)
college_names = [c[1] for c in COLLEGES]
college_validation = DataValidation(
    type="list",
    formula1=f'"{",".join(college_names)}"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle="كلية غير صحيحة",
    error="اختر من القائمة المنسدلة فقط. لا تكتب اسم كلية يدوياً.",
)
college_validation.add(f"A4:A1000")
ws2.add_data_validation(college_validation)

# Data Validation: قائمة التخصصات (العمود B) - منفصلة
specialty_names = [s[2] for s in SPECIALTIES]
specialty_validation = DataValidation(
    type="list",
    formula1=f'"{",".join(specialty_names)}"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle="تخصص غير صحيح",
    error="اختر من القائمة المنسدلة فقط.",
)
specialty_validation.add(f"B4:B1000")
ws2.add_data_validation(specialty_validation)

# Data Validation: المستوى (1-6) (العمود C)
level_validation = DataValidation(
    type="whole",
    operator="between",
    formula1=1,
    formula2=6,
    allow_blank=False,
    showErrorMessage=True,
    errorTitle="مستوى غير صحيح",
    error="المستوى يجب أن يكون رقماً من 1 إلى 6.",
)
level_validation.add(f"C4:C1000")
ws2.add_data_validation(level_validation)

# Data Validation: الفصل (1 أو 2) (العمود D)
semester_validation = DataValidation(
    type="list",
    formula1='"1,2"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle="فصل غير صحيح",
    error="الفصل يجب أن يكون 1 (الأول) أو 2 (الثاني).",
)
semester_validation.add(f"D4:D1000")
ws2.add_data_validation(semester_validation)

# Data Validation: نظري (نعم/لا) (العمود F)
yes_no_validation_1 = DataValidation(
    type="list",
    formula1='"نعم,لا"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle="قيمة غير صحيحة",
    error="اختر «نعم» أو «لا» فقط.",
)
yes_no_validation_1.add(f"F4:F1000")
ws2.add_data_validation(yes_no_validation_1)

# Data Validation: عملي (نعم/لا) (العمود G)
yes_no_validation_2 = DataValidation(
    type="list",
    formula1='"نعم,لا"',
    allow_blank=False,
    showErrorMessage=True,
    errorTitle="قيمة غير صحيحة",
    error="اختر «نعم» أو «لا» فقط.",
)
yes_no_validation_2.add(f"G4:G1000")
ws2.add_data_validation(yes_no_validation_2)

# تنسيق الصفوف الفارغة
for row_num in range(4, 60):
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws2[f"{col_letter}{row_num}"]
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell.font = Font(name=FONT_NAME, size=11, color=NEUTRAL_900)
        cell.border = Border(
            bottom=Side(style="thin", color=NEUTRAL_200)
        )
        if row_num % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=NEUTRAL_100)
        else:
            cell.fill = PatternFill("solid", fgColor=NEUTRAL_0)

# تجميد الصفوف العلوية
ws2.freeze_panes = 'A4'

# ============================================
# الورقة 3: الكليات_والتخصصات (مرجع)
# ============================================
ws3 = wb.create_sheet("🏛 الكليات_والتخصصات", 2)
ws3.sheet_view.rightToLeft = True

# العنوان
ws3.merge_cells('A1:E1')
ws3['A1'] = "🏛 الكليات والتخصصات (مرجع - لا تُعدّل)"
apply_style(ws3['A1'], title_style())
ws3.row_dimensions[1].height = 32

ws3.merge_cells('A2:E2')
ws3['A2'] = "هذه القائمة مرجعية. استخدم نفس الأسماء تماماً عند تعبئة ورقة المواد."
apply_style(ws3['A2'], subtitle_style())
ws3.row_dimensions[2].height = 22

# الترويسة
ws3_headers = [
    ("A", "الرقم", 10),
    ("B", "الكلية", 35),
    ("C", "التخصص", 40),
    ("D", "عدد المستويات", 18),
    ("E", "ملاحظة", 25),
]

for col_letter, header, width in ws3_headers:
    ws3.column_dimensions[col_letter].width = width
    cell = ws3[f"{col_letter}3"]
    cell.value = header
    apply_style(cell, column_header_style())

ws3.row_dimensions[3].height = 35

# البيانات
current_row = 4
for college_id, college_name, college_emoji in COLLEGES:
    college_specialties = [s for s in SPECIALTIES if s[1] == college_id]
    start_row = current_row
    for spec_id, _, spec_name, levels_count in college_specialties:
        ws3.cell(row=current_row, column=1, value=spec_id)
        ws3.cell(row=current_row, column=2, value=f"{college_emoji} {college_name}")
        ws3.cell(row=current_row, column=3, value=spec_name)
        ws3.cell(row=current_row, column=4, value=levels_count)
        ws3.cell(row=current_row, column=5, value=f"{levels_count} مستويات" if levels_count > 1 else "مستوى واحد")

        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            cell = ws3[f"{col_letter}{current_row}"]
            apply_style(cell, data_cell_style(odd_row=(current_row % 2 == 0)))

        ws3.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws3.cell(row=current_row, column=4).alignment = Alignment(horizontal="center", vertical="center")

        ws3.row_dimensions[current_row].height = 25
        current_row += 1

    if len(college_specialties) > 1:
        ws3.merge_cells(start_row=start_row, start_column=2, end_row=current_row - 1, end_column=2)

ws3.freeze_panes = 'A4'

# ============================================
# الورقة 4: مثال تقنية المعلومات
# ============================================
ws4 = wb.create_sheet("✅ مثال_تقنية_المعلومات", 3)
ws4.sheet_view.rightToLeft = True

# العنوان
ws4.merge_cells('A1:H1')
ws4['A1'] = "✅ مثال عملي: مواد تقنية معلومات (IT) - المستويان 1 و 2"
apply_style(ws4['A1'], title_style())
ws4.row_dimensions[1].height = 32

ws4.merge_cells('A2:H2')
ws4['A2'] = "هذا مثال على التعبئة الصحيحة. لا تُعدّله - استخدمه كمرجع فقط."
apply_style(ws4['A2'], subtitle_style())
ws4.row_dimensions[2].height = 22

# الترويسة (نفس ورقة المواد)
for col_letter, header, width in headers:
    ws4.column_dimensions[col_letter].width = width
    cell = ws4[f"{col_letter}3"]
    cell.value = header
    apply_style(cell, column_header_style())

ws4.row_dimensions[3].height = 35

# البيانات
it_college_name = "كلية الحاسبات وتكنولوجيا المعلومات"
it_specialty_name = "تقنية معلومات (IT)"

for i, (level, semester, subject_name, has_theory, has_practical) in enumerate(IT_SUBJECTS_EXAMPLE):
    row = 4 + i
    ws4.cell(row=row, column=1, value=it_college_name)
    ws4.cell(row=row, column=2, value=it_specialty_name)
    ws4.cell(row=row, column=3, value=level)
    ws4.cell(row=row, column=4, value=semester)
    ws4.cell(row=row, column=5, value=subject_name)
    ws4.cell(row=row, column=6, value=has_theory)
    ws4.cell(row=row, column=7, value=has_practical)
    ws4.cell(row=row, column=8, value="")

    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws4[f"{col_letter}{row}"]
        apply_style(cell, data_cell_style(odd_row=(row % 2 == 0)))

    for col in [3, 4, 6, 7]:
        ws4.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    ws4.row_dimensions[row].height = 25

# رسالة في نهاية المثال
end_row = 4 + len(IT_SUBJECTS_EXAMPLE) + 1
ws4.merge_cells(f'A{end_row}:H{end_row}')
ws4.cell(row=end_row, column=1, value="✅ هذا مثال صحيح للتعبئة. لاحظ: كل مادة في صف منفصل، والمستويات مرقمة 1-4، والفصول 1 أو 2.")
apply_style(ws4.cell(row=end_row, column=1), success_style())
ws4.row_dimensions[end_row].height = 30

ws4.freeze_panes = 'A4'

# ============================================
# حفظ الملف
# ============================================
OUTPUT_PATH = "/home/z/my-project/download/UST-قالب-المواد-الدراسية.xlsx"
wb.save(OUTPUT_PATH)

file_size_kb = os.path.getsize(OUTPUT_PATH) / 1024
print(f"✅ تم إنشاء قالب Excel بنجاح!")
print(f"📁 المسار: {OUTPUT_PATH}")
print(f"📊 الحجم: {file_size_kb:.2f} KB")
print(f"📋 الأوراق: 4 أوراق")
print(f"   1. 📋 الإرشادات - تعليمات التعبئة")
print(f"   2. 📝 المواد - القالب الرئيسي للتعبئة")
print(f"   3. 🏛 الكليات_والتخصصات - قائمة معبأة مسبقاً (34 تخصص)")
print(f"   4. ✅ مثال_تقنية_المعلومات - 20 مادة كنموذج")
print(f"")
print(f"✨ مميزات القالب:")
print(f"   • قوائم منسدلة للكليات والتخصصات (Data Validation)")
print(f"   • قوائم منسدلة لـ (نعم/لا) و (الفصل 1/2)")
print(f"   • تحقق من رقم المستوى (1-6)")
print(f"   • رسائل خطأ واضحة عند الإدخال الخاطئ")
print(f"   • تعليقات إرشادية على رؤوس الأعمدة")
print(f"   • اتجاه RTL (من اليمين لليسار)")
print(f"   • ألوان احترافية + صفوف متبادلة")
print(f"   • تجميد الصفوف العلوية للتنقّل السهل")
